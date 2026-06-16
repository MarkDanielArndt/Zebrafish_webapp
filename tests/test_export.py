import os


RESULTS = [
    {
        "filename": "fish1.png", "length": 1200.0, "curvature": 2,
        "ratio": 1.05, "eye_area": None, "eye_diameter": None, "error": None,
    },
    {
        "filename": "fish2.png", "length": None, "curvature": None,
        "ratio": None, "eye_area": None, "eye_diameter": None, "error": "Segmentation failed",
    },
]


def test_export_excel_creates_file(tmp_path):
    from zebrafish_analysis.slicer_extension.ZebrafishAnalysis.ZebrafishAnalysisLib.export import export_excel
    out = str(tmp_path / "results.xlsx")
    export_excel(RESULTS, out)
    assert os.path.exists(out)
    assert os.path.getsize(out) > 0


def test_export_csv_creates_file(tmp_path):
    from zebrafish_analysis.slicer_extension.ZebrafishAnalysis.ZebrafishAnalysisLib.export import export_csv
    out = str(tmp_path / "results.csv")
    export_csv(RESULTS, out)
    assert os.path.exists(out)
    lines = open(out).readlines()
    assert len(lines) == 3  # header + 2 rows


def test_export_excel_includes_error_column(tmp_path):
    import openpyxl
    from zebrafish_analysis.slicer_extension.ZebrafishAnalysis.ZebrafishAnalysisLib.export import export_excel
    out = str(tmp_path / "results.xlsx")
    export_excel(RESULTS, out)
    wb = openpyxl.load_workbook(out)
    headers = [cell.value for cell in wb.active[1]]
    assert "Error" in headers


def test_export_filtered_skips_excluded(tmp_path):
    from zebrafish_analysis.slicer_extension.ZebrafishAnalysis.ZebrafishAnalysisLib.export import export_excel
    import openpyxl
    out = str(tmp_path / "filtered.xlsx")
    filtered = [r for r in RESULTS if r["error"] is None]
    export_excel(filtered, out)
    wb = openpyxl.load_workbook(out)
    assert wb.active.max_row == 2  # header + 1 result row
